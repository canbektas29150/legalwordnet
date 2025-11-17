#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse, re, math
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook, Workbook

# ---------- normalize ----------
def normalize_tr(s):
    if s is None:
        return ""
    s = str(s).strip()
    # Şapkalı harfleri düz harflere indir
    s = (s
        .replace("Â", "A").replace("â", "a")
        .replace("Î", "İ").replace("î", "i")
        .replace("Û", "U").replace("û", "u")
        .replace("I", "ı").replace("İ", "i")
    )
    s = re.sub(r"\s+", " ", s)
    return s.casefold()

# ---------- POS belirleyici ----------
def guess_pos(defn: str) -> str:
    """Tanım sonundaki son kelime -mak/-mek ise VERB, aksi halde NOUN."""
    if not defn:
        return "NOUN"
    s = str(defn).strip().lower()
    words = re.findall(r"[A-Za-zÇĞİIÖŞÜçğıiöşüÂâÎîÛû]+", s)
    if words:
        last = words[-1]
        if last.endswith("mak") or last.endswith("mek"):
            return "VERB"
    return "NOUN"

# ---------- anlam tokenize ----------
def tokenize_def(text):
    """Tanımı kelime setine çevir (basit tokenizasyon)."""
    if not text:
        return set()
    s = str(text).lower()
    tokens = re.findall(r"[A-Za-zÇĞİIÖŞÜçğıiöşüÂâÎîÛû0-9]+", s)
    return set(tokens)

# ---------- benzerlik metotları ----------
def sim_overlap(new_tokens, old_tokens):
    """|A∩B| / |A|  (A = yeni tanım)"""
    if not new_tokens:
        return 0.0
    return len(new_tokens & old_tokens) / len(new_tokens)

def sim_jaccard(tokens_a, tokens_b):
    """|A∩B| / |A∪B|"""
    if not tokens_a and not tokens_b:
        return 0.0
    inter = len(tokens_a & tokens_b)
    union = len(tokens_a | tokens_b)
    if union == 0:
        return 0.0
    return inter / union

def sim_tfidf_cosine(tokens_a, tokens_b, df_counter, doc_count):
    """
    Basit TF-IDF + cosine:
    - tf = 1 (sadece var/yok)
    - idf = log((N+1)/(df+1)) + 1
    """
    if not tokens_a or not tokens_b or doc_count == 0:
        return 0.0

    vocab = tokens_a | tokens_b
    num = 0.0
    sum_a = 0.0
    sum_b = 0.0

    for t in vocab:
        df = df_counter.get(t, 0)
        idf = math.log((doc_count + 1) / (df + 1)) + 1.0
        wa = idf if t in tokens_a else 0.0
        wb = idf if t in tokens_b else 0.0
        num += wa * wb
        sum_a += wa * wa
        sum_b += wb * wb

    denom = (sum_a ** 0.5) * (sum_b ** 0.5)
    if denom == 0.0:
        return 0.0
    return num / denom

# ---------- header bulucu ----------
def find_col(ws, wanted, search_rows=10):
    """
    ws içinde istenen başlık adlarını bulup 1-bazlı kolon indekslerini döndürür.
    'DEFINITION' yerine 'ANLAM' da gelebilir, bu yüzden eş adları destekliyoruz.
    """
    synonyms = {
        "DEFINITION": {"DEFINITION", "ANLAM"},
        "KELİME": {"KELİME", "KELIME", "WORD"},
        "POS": {"POS"},
        "R": {"R"},
        "ID": {"ID"},
        "EXAMPLE SENTENCE": {"EXAMPLE SENTENCE", "EXAMPLE", "ÖRNEK"},
    }

    wanted_sets = {w: {w.lower()} | {x.lower() for x in synonyms.get(w, {w})} for w in wanted}
    found = {w: None for w in wanted}

    for row in ws.iter_rows(min_row=1, max_row=search_rows, values_only=True):
        if not row:
            continue
        for j, val in enumerate(row, start=1):
            if val is None:
                continue
            key = str(val).strip().lower()
            for w in wanted:
                if found[w] is None and key in wanted_sets[w]:
                    found[w] = j
        if all(found.get(w) for w in wanted if w != "EXAMPLE SENTENCE"):
            break
    return found

# ---------- yeni sözlük kelimelerini oku ----------
def read_new_words(new_path):
    wb = load_workbook(new_path, read_only=True, data_only=True)
    skip = {"özet", "toplam"}
    data = {}
    for sh in wb.sheetnames:
        if sh.lower() in skip:
            continue
        ws = wb[sh]
        idx = find_col(ws, ["kelime", "anlam"])
        kcol, acol = idx.get("kelime"), idx.get("anlam")
        if not kcol:
            continue
        pairs = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            k = r[kcol - 1] if len(r) >= kcol else None
            a = r[acol - 1] if acol and len(r) >= acol else ""
            if k:
                pairs.append((k, a))
        data[sh] = pairs
    return data

# ---------- güvenli başlık oluşturucu ----------
NEEDED_HEADERS = ["R", "KELİME", "ID", "POS", "DEFINITION", "EXAMPLE SENTENCE"]

def ensure_headers(ws):
    if ws.max_row < 1:
        ws.append(NEEDED_HEADERS)
    header_row = 1
    header_vals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    if not any(header_vals):
        for i, name in enumerate(NEEDED_HEADERS, start=1):
            ws.cell(header_row, i, name)

    idx = find_col(ws, NEEDED_HEADERS)
    colmap = {}
    for name in NEEDED_HEADERS:
        c = idx.get(name)
        if not c:
            c = ws.max_column + 1
            ws.cell(1, c, name)
        colmap[name] = c
    return colmap

# ---------- ana işlem ----------
def update_and_flag(old_path, new_path, out_path, sim_threshold=0.5, sim_method="jaccard"):
    wb_old = load_workbook(old_path, read_only=False, data_only=True)
    new_data = read_new_words(new_path)

    stats = {}  # { sheet_name: {"total":0, "matched":0, "added":0} }
    threshold_matches = []  # log

    # tf-idf için global df ve doküman sayısı
    df_counter = defaultdict(int)
    doc_count = 0

    # cache: sheet_name -> (ws, colmap, old_norm_map)
    # old_norm_map: { norm : [ {"row":int, "def":str, "tokens":set}, ... ] }
    cache = {}

    def ensure_page(letter):
        nonlocal doc_count
        if letter in cache:
            return cache[letter]

        if letter not in wb_old.sheetnames:
            ws = wb_old.create_sheet(letter)
        else:
            ws = wb_old[letter]

        colmap = ensure_headers(ws)
        kcol = colmap["KELİME"]
        dcol = colmap["DEFINITION"]

        old_norm_map = {}

        for row in ws.iter_rows(min_row=2, values_only=False):
            cell_val = row[kcol - 1].value
            if not cell_val:
                continue
            norm = normalize_tr(cell_val)
            def_val = row[dcol - 1].value if dcol <= len(row) else ""
            tokens = tokenize_def(def_val)

            if tokens:
                doc_count += 1
                for tok in tokens:
                    df_counter[tok] += 1

            entry = {
                "row": row[0].row,
                "def": def_val,
                "tokens": tokens,
            }
            old_norm_map.setdefault(norm, []).append(entry)

        cache[letter] = (ws, colmap, old_norm_map)
        return cache[letter]

    print(f"Kullanılan benzerlik threshold'u: {sim_threshold}")
    print(f"Kullanılan benzerlik metodu: {sim_method}\n")

    # yeni veriyi tara
    for sh, pairs in new_data.items():
        for (kelime, anlam) in pairs:
            norm = normalize_tr(kelime)
            if not norm:
                continue

            first = kelime.strip()[0] if kelime.strip() else ""
            if   first in ("Â","â"): target = "A"
            elif first in ("Î","î"): target = "İ"
            elif first in ("Û","û"): target = "U"
            else:                    target = sh

            ws, colmap, old_norm_map = ensure_page(target)

            if target not in stats:
                stats[target] = {"total": 0, "matched": 0, "added": 0}

            stats[target]["total"] += 1
            kcol = colmap["KELİME"]
            dcol = colmap["DEFINITION"]
            pcol = colmap["POS"]
            rcol = colmap["R"]

            if norm in old_norm_map:
                candidates = old_norm_map[norm]

                # tek satır varsa direkt match (score=1)
                if len(candidates) == 1:
                    row_idx = candidates[0]["row"]
                    ws.cell(row=row_idx, column=rcol, value=1)
                    stats[target]["matched"] += 1
                    threshold_matches.append({
                        "sheet": target,
                        "word": kelime,
                        "mode": "single",
                        "row": row_idx,
                        "score": 1.0,
                        "new_def": anlam,
                        "old_def": candidates[0]["def"],
                        "candidate_count": 1,
                    })
                else:
                    new_tokens = tokenize_def(anlam)
                    best_row = None
                    best_score = 0.0
                    best_old_def = None

                    if new_tokens:
                        for cand in candidates:
                            old_tokens = cand["tokens"]

                            if sim_method == "tfidf":
                                score = sim_tfidf_cosine(new_tokens, old_tokens, df_counter, doc_count)
                            elif sim_method == "jaccard":
                                score = sim_jaccard(new_tokens, old_tokens)
                            else:
                                score = sim_overlap(new_tokens, old_tokens)

                            if score > best_score:
                                best_score = score
                                best_row = cand["row"]
                                best_old_def = cand["def"]

                    if best_row is not None and best_score >= sim_threshold:
                        ws.cell(row=best_row, column=rcol, value=1)
                        stats[target]["matched"] += 1
                        threshold_matches.append({
                            "sheet": target,
                            "word": kelime,
                            "mode": f"duplicate+{sim_method}",
                            "row": best_row,
                            "score": best_score,
                            "new_def": anlam,
                            "old_def": best_old_def,
                            "candidate_count": len(candidates),
                        })
                    else:
                        stats[target]["added"] += 1
                        new_row_idx = ws.max_row + 1
                        ws.cell(new_row_idx, kcol, kelime)
                        ws.cell(new_row_idx, dcol, anlam)
                        ws.cell(new_row_idx, pcol, guess_pos(anlam))
                        ws.cell(new_row_idx, rcol, 0)
                        old_norm_map.setdefault(norm, []).append({
                            "row": new_row_idx,
                            "def": anlam,
                            "tokens": tokenize_def(anlam),
                        })
            else:
                stats[target]["added"] += 1
                new_row_idx = ws.max_row + 1
                ws.cell(new_row_idx, kcol, kelime)
                ws.cell(new_row_idx, dcol, anlam)
                ws.cell(new_row_idx, pcol, guess_pos(anlam))
                ws.cell(new_row_idx, rcol, 0)
                old_norm_map.setdefault(norm, []).append({
                    "row": new_row_idx,
                    "def": anlam,
                    "tokens": tokenize_def(anlam),
                })

    # --- R boşsa 0 yap ---
    for sh in wb_old.sheetnames:
        ws = wb_old[sh]
        idx = find_col(ws, ["R"])
        rcol = idx.get("R")
        if not rcol:
            continue
        for row in ws.iter_rows(min_row=2):
            cell = row[rcol - 1]
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = 0

    wb_old.save(out_path)

    # ----- threshold ile eşleşenlerin çıktısı -----
    print("\n==== THRESHOLD İLE EŞLEŞEN SATIRLAR ====")
    if not threshold_matches:
        print("Bu threshold ile hiç eşleşme yapılmadı.")
    else:
        for m in threshold_matches:
            print(
                f"[{m['sheet']}] kelime='{m['word']}' "
                f"(mode={m['mode']}, satır={m['row']}, skor={m['score']:.3f}, "
                f"aday_sayısı={m['candidate_count']})"
            )
            # debug istersen aç:
            # print(f"  NEW: {m['new_def']}")
            # print(f"  OLD: {m['old_def']}")
        print(f"Toplam threshold-match sayısı: {len(threshold_matches)}")

    # ----- çok adaylı match'ler için ayrı Excel -----
    ambiguous = [m for m in threshold_matches if m.get("candidate_count", 0) > 1]

    if ambiguous:
        amb_wb = Workbook()
        ws_amb = amb_wb.active
        ws_amb.title = "AmbiguousMatches"
        ws_amb.append([
            "Sheet",
            "Word",
            "Score",
            "CandidateCount",
            "MatchedRow",
            "NewDefinition",
            "OldDefinition",
            "Method",
        ])

        for m in ambiguous:
            ws_amb.append([
                m["sheet"],
                m["word"],
                round(m["score"], 3),
                m["candidate_count"],
                m["row"],
                m["new_def"],
                m["old_def"],
                m["mode"],
            ])

        amb_path = Path(out_path)
        amb_file = amb_path.with_name(amb_path.stem + "_ambiguous.xlsx")
        amb_wb.save(amb_file)
        print(f"\n✔ {len(ambiguous)} adet çok adaylı eşleşme ayrı dosyaya yazıldı: {amb_file}")
    else:
        print("\nÇok adaylı (candidate_count>1) eşleşme bulunmadı, ek Excel üretilmedi.")

    # ----- sayfa bazlı özet -----
    print("\n==== SAYFA BAZLI ÖZET ====")
    total_all = sum(d["total"] for d in stats.values())
    matched_all = sum(d["matched"] for d in stats.values())
    added_all = sum(d["added"] for d in stats.values())

    for letter in sorted(stats.keys()):
        d = stats[letter]
        print(f"[{letter}] toplam={d['total']}, eşleşen(R=1)={d['matched']}, eklenen(R=0)={d['added']}")

    print("---- GENEL ÖZET ----")
    print(f"Toplam kelime: {total_all}")
    print(f"Eşleşen (R=1): {matched_all}")
    print(f"Yeni eklenen (R=0): {added_all}")
    print(f"✔ Güncellendi ve kaydedildi: {out_path}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--old", required=True)   # HukukSözlüğü.xlsx
    ap.add_argument("--new", required=True)   # sozlukafull.xlsx
    ap.add_argument("--out", default="updated_flagged.xlsx")
    args = ap.parse_args()

    # Threshold
    try:
        raw = input("Benzerlik threshold (0-1 arası, boş bırakılırsa 0.5): ").strip()
        if raw == "":
            sim_thr = 0.5
        else:
            sim_thr = float(raw)
            if sim_thr < 0 or sim_thr > 1:
                print("Geçersiz değer, 0.5 kullanılacak.")
                sim_thr = 0.5
    except Exception:
        print("Threshold okunamadı, 0.5 kullanılacak.")
        sim_thr = 0.5

    # Metot seçimi
    print("Benzerlik metodu seç:")
    print("  1 = Overlap (|A∩B| / |A|)")
    print("  2 = Jaccard (|A∩B| / |A∪B|) [varsayılan]")
    print("  3 = TF-IDF + Cosine")
    m_raw = input("Seçimin (1/2/3, boş bırakılırsa 2): ").strip()

    if m_raw == "1":
        sim_method = "overlap"
    elif m_raw == "3":
        sim_method = "tfidf"
    else:
        sim_method = "jaccard"

    update_and_flag(args.old, args.new, args.out,
                    sim_threshold=sim_thr,
                    sim_method=sim_method)

if __name__ == "__main__":
    import sys
    if len(sys.argv) == 1:
        sys.argv = [
            "x",
            "--old", "HukukSözlüğü.xlsx",
            "--new", "sozlukafull.xlsx",
            "--out", "sozluk_a_flagged_yeni.xlsx",
        ]
    main()
