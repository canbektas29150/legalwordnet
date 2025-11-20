#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse, re, math
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook, Workbook

# ---------- normalize ----------
def normalize_tr(s):
    if s is None:
        return ""
    s = str(s).strip()
    # ÅapkalÄ± harfleri dÃ¼z harflere indir
    s = (s
        .replace("Ã‚", "A").replace("Ã¢", "a")
        .replace("Ã", "Ä°").replace("Ã®", "i")
        .replace("Ã›", "U").replace("Ã»", "u")
        .replace("I", "Ä±").replace("Ä°", "i")
    )
    s = re.sub(r"\s+", " ", s)
    return s.casefold()

# ---------- POS belirleyici ----------
def guess_pos(defn: str) -> str:
    """TanÄ±m sonundaki son kelime -mak/-mek ise VERB, aksi halde NOUN."""
    if not defn:
        return "NOUN"
    s = str(defn).strip().lower()
    words = re.findall(r"[A-Za-zÃ‡ÄÄ°IÃ–ÅÃœÃ§ÄŸÄ±iÃ¶ÅŸÃ¼Ã‚Ã¢ÃÃ®Ã›Ã»]+", s)
    if words:
        last = words[-1]
        if last.endswith("mak") or last.endswith("mek"):
            return "VERB"
    return "NOUN"

# ---------- anlam tokenize ----------
def tokenize_def(text):
    """TanÄ±mÄ± kelime setine Ã§evir (basit tokenizasyon)."""
    if not text:
        return set()
    s = str(text).lower()
    tokens = re.findall(r"[A-Za-zÃ‡ÄÄ°IÃ–ÅÃœÃ§ÄŸÄ±iÃ¶ÅŸÃ¼Ã‚Ã¢ÃÃ®Ã›Ã»0-9]+", s)
    return set(tokens)

# ---------- benzerlik metotlarÄ± ----------
def sim_overlap(new_tokens, old_tokens):
    """|Aâˆ©B| / |A|  (A = yeni tanÄ±m)"""
    if not new_tokens:
        return 0.0
    return len(new_tokens & old_tokens) / len(new_tokens)

def sim_jaccard(tokens_a, tokens_b):
    """|Aâˆ©B| / |AâˆªB|"""
    if not tokens_a and not tokens_b:
        return 0.0
    inter = len(tokens_a & tokens_b)
    union = len(tokens_a | tokens_b)
    if union == 0:
        return 0.0
    return inter / union

def sim_tfidf_cosine(tokens_a, tokens_b, df_counter, doc_count):
    """
    Basit TF-IDF + cosine:
    - tf = 1 (sadece var/yok)
    - idf = log((N+1)/(df+1)) + 1
    """
    if not tokens_a or not tokens_b or doc_count == 0:
        return 0.0

    vocab = tokens_a | tokens_b
    num = 0.0
    sum_a = 0.0
    sum_b = 0.0

    for t in vocab:
        df = df_counter.get(t, 0)
        idf = math.log((doc_count + 1) / (df + 1)) + 1.0
        wa = idf if t in tokens_a else 0.0
        wb = idf if t in tokens_b else 0.0
        num += wa * wb
        sum_a += wa * wa
        sum_b += wb * wb

    denom = (sum_a ** 0.5) * (sum_b ** 0.5)
    if denom == 0.0:
        return 0.0
    return num / denom

# ---------- header bulucu ----------
def find_col(ws, wanted, search_rows=10):
    """
    ws iÃ§inde istenen baÅŸlÄ±k adlarÄ±nÄ± bulup 1-bazlÄ± kolon indekslerini dÃ¶ndÃ¼rÃ¼r.
    'DEFINITION' yerine 'ANLAM' da gelebilir, bu yÃ¼zden eÅŸ adlarÄ± destekliyoruz.
    """
    synonyms = {
        "DEFINITION": {"DEFINITION", "ANLAM"},
        "KELÄ°ME": {"KELÄ°ME", "KELIME", "WORD"},
        "POS": {"POS"},
        "R": {"R"},
        "ID": {"ID"},
        "EXAMPLE SENTENCE": {"EXAMPLE SENTENCE", "EXAMPLE", "Ã–RNEK"},
    }

    wanted_sets = {w: {w.lower()} | {x.lower() for x in synonyms.get(w, {w})} for w in wanted}
    found = {w: None for w in wanted}

    for row in ws.iter_rows(min_row=1, max_row=search_rows, values_only=True):
        if not row:
            continue
        for j, val in enumerate(row, start=1):
            if val is None:
                continue
            key = str(val).strip().lower()
            for w in wanted:
                if found[w] is None and key in wanted_sets[w]:
                    found[w] = j
        if all(found.get(w) for w in wanted if w != "EXAMPLE SENTENCE"):
            break
    return found

# ---------- yeni sÃ¶zlÃ¼k kelimelerini oku ----------
def read_new_words(new_path):
    wb = load_workbook(new_path, read_only=True, data_only=True)
    skip = {"Ã¶zet", "toplam"}
    data = {}
    for sh in wb.sheetnames:
        if sh.lower() in skip:
            continue
        ws = wb[sh]
        idx = find_col(ws, ["kelime", "anlam"])
        kcol, acol = idx.get("kelime"), idx.get("anlam")
        if not kcol:
            continue
        pairs = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            k = r[kcol - 1] if len(r) >= kcol else None
            a = r[acol - 1] if acol and len(r) >= acol else ""
            if k:
                pairs.append((k, a))
        data[sh] = pairs
    return data

# ---------- gÃ¼venli baÅŸlÄ±k oluÅŸturucu ----------
NEEDED_HEADERS = ["R", "KELÄ°ME", "ID", "POS", "DEFINITION", "EXAMPLE SENTENCE"]

def ensure_headers(ws):
    if ws.max_row < 1:
        ws.append(NEEDED_HEADERS)
    header_row = 1
    header_vals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    if not any(header_vals):
        for i, name in enumerate(NEEDED_HEADERS, start=1):
            ws.cell(header_row, i, name)

    idx = find_col(ws, NEEDED_HEADERS)
    colmap = {}
    for name in NEEDED_HEADERS:
        c = idx.get(name)
        if not c:
            c = ws.max_column + 1
            ws.cell(1, c, name)
        colmap[name] = c
    return colmap

# ---------- ana iÅŸlem ----------
def update_and_flag(old_path, new_path, out_path, sim_threshold=0.5, sim_method="jaccard"):
    wb_old = load_workbook(old_path, read_only=False, data_only=True)
    new_data = read_new_words(new_path)

    stats = {}              # { sheet_name: {"total":0, "matched":0, "added":0} }
    threshold_matches = []  # sadece seÃ§ilen match'ler (log iÃ§in)
    ambiguous_rows = []     # candidate_count>1 ve match olan tÃ¼m adaylar (hocanÄ±n sÃ¶zlÃ¼ÄŸÃ¼ tarafÄ±)

    # tf-idf iÃ§in global df ve dokÃ¼man sayÄ±sÄ±
    df_counter = defaultdict(int)
    doc_count = 0

    # cache: sheet_name -> (ws, colmap, old_norm_map)
    # old_norm_map: { norm : [ {"row":int, "def":str, "tokens":set}, ... ] }
    cache = {}

    def ensure_page(letter):
        nonlocal doc_count
        if letter in cache:
            return cache[letter]

        if letter not in wb_old.sheetnames:
            ws = wb_old.create_sheet(letter)
        else:
            ws = wb_old[letter]

        colmap = ensure_headers(ws)
        kcol = colmap["KELÄ°ME"]
        dcol = colmap["DEFINITION"]

        old_norm_map = {}

        for row in ws.iter_rows(min_row=2, values_only=False):
            cell_val = row[kcol - 1].value
            if not cell_val:
                continue
            norm = normalize_tr(cell_val)
            def_val = row[dcol - 1].value if dcol <= len(row) else ""
            tokens = tokenize_def(def_val)

            if tokens:
                doc_count += 1
                for tok in tokens:
                    df_counter[tok] += 1

            entry = {
                "row": row[0].row,
                "def": def_val,
                "tokens": tokens,
            }
            old_norm_map.setdefault(norm, []).append(entry)

        cache[letter] = (ws, colmap, old_norm_map)
        return cache[letter]

    print(f"KullanÄ±lan benzerlik threshold'u: {sim_threshold}")
    print(f"KullanÄ±lan benzerlik metodu: {sim_method}\n")

    # yeni veriyi tara
    for sh, pairs in new_data.items():
        for (kelime, anlam) in pairs:
            norm = normalize_tr(kelime)
            if not norm:
                continue

            first = kelime.strip()[0] if kelime.strip() else ""
            if   first in ("Ã‚","Ã¢"): target = "A"
            elif first in ("Ã","Ã®"): target = "Ä°"
            elif first in ("Ã›","Ã»"): target = "U"
            else:                    target = sh

            ws, colmap, old_norm_map = ensure_page(target)

            if target not in stats:
                stats[target] = {"total": 0, "matched": 0, "added": 0}

            stats[target]["total"] += 1
            kcol = colmap["KELÄ°ME"]
            dcol = colmap["DEFINITION"]
            pcol = colmap["POS"]
            rcol = colmap["R"]

            if norm in old_norm_map:
                candidates = old_norm_map[norm]

                # ---- 1) Tek aday varsa: her zamanki gibi R=1 yaz ----
                if len(candidates) == 1:
                    row_idx = candidates[0]["row"]
                    ws.cell(row=row_idx, column=rcol, value=1)
                    stats[target]["matched"] += 1
                    threshold_matches.append({
                        "sheet": target,
                        "word": kelime,
                        "mode": "single",
                        "row": row_idx,
                        "score": 1.0,
                        "new_def": anlam,
                        "old_def": candidates[0]["def"],
                        "candidate_count": 1,
                    })

                # ---- 2) Birden fazla aday varsa: R'e HÄ°Ã‡BÄ°R ÅEY YAZMA ----
                else:
                    new_tokens = tokenize_def(anlam)
                    best_row = None
                    best_score = 0.0
                    best_old_def = None
                    cand_scores = []  # her candidate iÃ§in (entry, score)

                    if new_tokens:
                        for cand in candidates:
                            old_tokens = cand["tokens"]

                            if sim_method == "tfidf":
                                score = sim_tfidf_cosine(new_tokens, old_tokens, df_counter, doc_count)
                            elif sim_method == "jaccard":
                                score = sim_jaccard(new_tokens, old_tokens)
                            else:
                                score = sim_overlap(new_tokens, old_tokens)

                            cand_scores.append((cand, score))

                            if score > best_score:
                                best_score = score
                                best_row = cand["row"]
                                best_old_def = cand["def"]

                    # EÄŸer threshold Ã¼stÃ¼nde iyi bir eÅŸleÅŸme varsa:
                    # - ArtÄ±k R=1 YAZMIYORUZ
                    # - Sadece ambiguous loguna ekliyoruz
                    if best_row is not None and best_score >= sim_threshold:
                        # R'e dokunmuyoruz, stats[target]["matched"] da artÄ±rmÄ±yoruz
                        # threshold_matches'e de eklemiyoruz ki konsolda "match" gÃ¶rÃ¼nmesin

                        # Ambiguous Excel iÃ§in: tÃ¼m adaylar + score + chosen flag
                        for cand, score in cand_scores:
                            ambiguous_rows.append({
                                "sheet": target,
                                "word": kelime,
                                "row": cand["row"],          # HukukSÃ¶zlÃ¼ÄŸÃ¼ satÄ±rÄ±
                                "score": score,
                                "chosen": (cand["row"] == best_row),  # sadece Ã¶neri olarak
                                "candidate_count": len(candidates),
                                "new_def": anlam,
                                "old_def": cand["def"],
                                "method": sim_method,
                            })
                    else:
                        # Threshold altÄ±nda â†’ "hiÃ§ eÅŸleÅŸmeyen" say, YENÄ° SATIR EKLE
                        stats[target]["added"] += 1
                        new_row_idx = ws.max_row + 1
                        ws.cell(new_row_idx, kcol, kelime)
                        ws.cell(new_row_idx, dcol, anlam)
                        ws.cell(new_row_idx, pcol, guess_pos(anlam))
                        ws.cell(new_row_idx, rcol, 1)   # ğŸ”´ ESKÄ°DE 0'DÄ±, ARTIK 1
                        old_norm_map.setdefault(norm, []).append({
                            "row": new_row_idx,
                            "def": anlam,
                            "tokens": tokenize_def(anlam),
                        })

            else:
                # ---- 3) HiÃ§ aday yoksa: direkt yeni satÄ±r, R=1 ----
                stats[target]["added"] += 1
                new_row_idx = ws.max_row + 1
                ws.cell(new_row_idx, kcol, kelime)
                ws.cell(new_row_idx, dcol, anlam)
                ws.cell(new_row_idx, pcol, guess_pos(anlam))
                ws.cell(new_row_idx, rcol, 1)   # ğŸ”´ ESKÄ°DE 0'DÄ±, ARTIK 1
                old_norm_map.setdefault(norm, []).append({
                    "row": new_row_idx,
                    "def": anlam,
                    "tokens": tokenize_def(anlam),
                })


    # --- R boÅŸsa 0 yap ---
    for sh in wb_old.sheetnames:
        ws = wb_old[sh]
        idx = find_col(ws, ["R"])
        rcol = idx.get("R")
        if not rcol:
            continue
        for row in ws.iter_rows(min_row=2):
            cell = row[rcol - 1]
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = 0

    wb_old.save(out_path)

    # ----- threshold ile eÅŸleÅŸenlerin konsol Ã§Ä±ktÄ±sÄ± -----
    print("\n==== THRESHOLD Ä°LE EÅLEÅEN SATIRLAR ====")
    if not threshold_matches:
        print("Bu threshold ile hiÃ§ eÅŸleÅŸme yapÄ±lmadÄ±.")
    else:
        for m in threshold_matches:
            print(
                f"[{m['sheet']}] kelime='{m['word']}' "
                f"(mode={m['mode']}, satÄ±r={m['row']}, skor={m['score']:.3f}, "
                f"aday_sayÄ±sÄ±={m['candidate_count']})"
            )
        print(f"Toplam threshold-match sayÄ±sÄ±: {len(threshold_matches)}")

    # ----- Ã§ok adaylÄ± match'ler iÃ§in ayrÄ± Excel (hocanÄ±n sÃ¶zlÃ¼ÄŸÃ¼ tarafÄ±, her aday satÄ±r + score) -----
    # ambiguous_rows zaten sadece: len(candidates)>1 VE best_score>=threshold durumunda doluyor.
    if ambiguous_rows:
        amb_wb = Workbook()
        ws_amb = amb_wb.active
        ws_amb.title = "AmbiguousMatches"
        ws_amb.append([
            "Sheet",
            "Word",
            "OldRow",          # HukukSÃ¶zlÃ¼ÄŸÃ¼ satÄ±r numarasÄ±
            "Score",
            "Chosen",          # Bu satÄ±r mÄ± R=1 aldÄ±?
            "CandidateCount",
            "NewDefinition",
            "OldDefinition",
            "Method",
        ])

        for m in ambiguous_rows:
            ws_amb.append([
                m["sheet"],
                m["word"],
                m["row"],
                round(m["score"], 3),
                1 if m["chosen"] else 0,
                m["candidate_count"],
                m["new_def"],
                m["old_def"],
                m["method"],
            ])

        amb_path = Path(out_path)
        amb_file = amb_path.with_name(amb_path.stem + "_ambiguous.xlsx")
        amb_wb.save(amb_file)
        print(f"\nâœ” {len(ambiguous_rows)} aday satÄ±r (candidate>1 & matched) ayrÄ± dosyaya yazÄ±ldÄ±: {amb_file}")
    else:
        print("\nÃ‡ok adaylÄ± ve match edilmiÅŸ eÅŸleÅŸme bulunmadÄ±, ek Excel Ã¼retilmedi.")

    # ----- sayfa bazlÄ± Ã¶zet -----
    print("\n==== SAYFA BAZLI Ã–ZET ====")
    total_all = sum(d["total"] for d in stats.values())
    matched_all = sum(d["matched"] for d in stats.values())
    added_all = sum(d["added"] for d in stats.values())

    for letter in sorted(stats.keys()):
        d = stats[letter]
        print(f"[{letter}] toplam={d['total']}, eÅŸleÅŸen(R=1)={d['matched']}, eklenen(R=0)={d['added']}")

    print("---- GENEL Ã–ZET ----")
    print(f"Toplam kelime: {total_all}")
    print(f"EÅŸleÅŸen (R=1): {matched_all}")
    print(f"Yeni eklenen (R=0): {added_all}")
    print(f"âœ” GÃ¼ncellendi ve kaydedildi: {out_path}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--old", required=True)   # HukukSÃ¶zlÃ¼ÄŸÃ¼.xlsx
    ap.add_argument("--new", required=True)   # sozlukafull.xlsx
    ap.add_argument("--out", default="updated_flagged.xlsx")
    args = ap.parse_args()

    # Threshold
    try:
        raw = input("Benzerlik threshold (0-1 arasÄ±, boÅŸ bÄ±rakÄ±lÄ±rsa 0.5): ").strip()
        if raw == "":
            sim_thr = 0.5
        else:
            sim_thr = float(raw)
            if sim_thr < 0 or sim_thr > 1:
                print("GeÃ§ersiz deÄŸer, 0.5 kullanÄ±lacak.")
                sim_thr = 0.5
    except Exception:
        print("Threshold okunamadÄ±, 0.5 kullanÄ±lacak.")
        sim_thr = 0.5

    # Metot seÃ§imi
    print("Benzerlik metodu seÃ§:")
    print("  1 = Overlap (|Aâˆ©B| / |A|)")
    print("  2 = Jaccard (|Aâˆ©B| / |AâˆªB|) [varsayÄ±lan]")
    print("  3 = TF-IDF + Cosine")
    m_raw = input("SeÃ§imin (1/2/3, boÅŸ bÄ±rakÄ±lÄ±rsa 2): ").strip()

    if m_raw == "1":
        sim_method = "overlap"
    elif m_raw == "3":
        sim_method = "tfidf"
    else:
        sim_method = "jaccard"

    update_and_flag(args.old, args.new, args.out,
                    sim_threshold=sim_thr,
                    sim_method=sim_method)

if __name__ == "__main__":
    import sys
    if len(sys.argv) == 1:
        sys.argv = [
            "x",
            "--old", "HukukSÃ¶zlÃ¼ÄŸÃ¼.xlsx",
            "--new", "sozlukÃ§yeniteseract.xlsx",
            "--out", "sozlukÃ§compared.xlsx",
        ]
    main()
